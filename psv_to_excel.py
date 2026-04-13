"""
shuffler/psv_to_excel.py — Convert a ♔-delimited PSV file to an Excel .xlsx file.

Usage
-----
    from shuffler.psv_to_excel import psv_to_excel
    psv_to_excel("path/to/file.psv")

Or from the command line:
    python -m shuffler.psv_to_excel path/to/file.psv

The output file is written to the same directory with the same stem but the
.xlsx extension (e.g., combined.psv → combined.xlsx).

Issues handled
--------------
1. Field separator  : ♔ is the column delimiter; handled by psv_io.read_psv.
2. In-field newlines: ⏎ is decoded back to \\n by read_psv, then written into
   the Excel cell as a proper in-cell line break (Alt+Enter equivalent).
   wrap_text is enabled so the line breaks render.  Each PSV record becomes
   exactly one Excel row regardless of how many logical lines are in a field.
3. Unicode fidelity : openpyxl writes a true .xlsx (UTF-8 XML internally), so
   curly quotes, em-dashes, and other non-ASCII characters are stored and
   displayed correctly — no Windows-1252 / Latin-1 mis-encoding.
4. Cell-length limit: Excel silently corrupts cells longer than 32 767
   characters.  Any value that exceeds this limit is truncated to 32 764
   characters and "…" is appended as a signal that truncation occurred.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

from shuffler.psv_io import read_psv

# Excel's hard limit on characters per cell
_EXCEL_MAX_CELL = 32_767
_TRUNCATION_MARKER = "…"
_TRUNCATION_LIMIT = _EXCEL_MAX_CELL - len(_TRUNCATION_MARKER)

# Characters illegal in XML 1.0 (as used by lxml / openpyxl's xlsx writer):
#   U+0000–U+0008  control chars
#   U+000B–U+000C  vertical tab, form feed
#   U+000E–U+001F  more control chars
#   U+D800–U+DFFF  UTF-16 surrogate halves (not valid Unicode scalar values)
#   U+FFFE–U+FFFF  Unicode non-characters
# Keeps U+0009 (tab), U+000A (LF), U+000D (CR) and all printable characters.
_ILLEGAL_CHARS_RE = re.compile(
    "[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF\uFFFE\uFFFF]"
)


def _sanitize(value: str) -> str:
    """
    Prepare a field value for writing to an Excel cell.

    * Strips control characters that are illegal in XML 1.0 / xlsx
      (keeps tab \\t, newline \\n, and carriage return \\r which are valid).
    * Truncates values that exceed Excel's 32 767-character cell limit.
    """
    value = _ILLEGAL_CHARS_RE.sub("", value)
    if len(value) > _EXCEL_MAX_CELL:
        value = value[:_TRUNCATION_LIMIT] + _TRUNCATION_MARKER
    return value


def psv_to_excel(file_path: str | Path) -> Path:
    """
    Load a ♔-delimited PSV file and write an equivalent .xlsx file.

    Parameters
    ----------
    file_path:
        Path to the source .psv file.

    Returns
    -------
    Path
        The path of the newly created .xlsx file.
    """
    src = Path(file_path)
    dst = src.with_suffix(".xlsx")

    rows, fieldnames = read_psv(src)

    wb = openpyxl.Workbook()
    ws = wb.active

    # ── Header row ────────────────────────────────────────────────────────────
    ws.append([_sanitize(h) for h in fieldnames])

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row in rows:
        cell_values = [_sanitize(str(row.get(col, "") or "")) for col in fieldnames]
        ws.append(cell_values)

        # Enable wrap_text on any cell that contains an in-cell line break so
        # the newlines actually render in Excel rather than appearing as boxes.
        excel_row = ws.max_row
        for col_idx, val in enumerate(cell_values, start=1):
            if "\n" in val:
                ws.cell(row=excel_row, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    wb.save(dst)
    print(f"✓  Saved {dst}")
    return dst


# ── CLI entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m shuffler.psv_to_excel <file.psv>", file=sys.stderr)
        sys.exit(1)
    psv_to_excel(sys.argv[1])
